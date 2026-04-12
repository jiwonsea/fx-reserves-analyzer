"""openpyxl Excel 리포트 — 데이터 시트 + Dashboard 통합 차트.

시트 구성:
  Dashboard      — 차트 5개 + 설명 한 곳에 모음 (첫 번째 시트)
  Summary        — 핵심 통계 요약
  TimeSeries     — 월별 외환보유고/환율 원시 데이터
  DeltaReturn    — MoM 변화량/변동률 데이터
  Granger        — lag × 방향 p-value 표
  IRF            — 충격반응함수 데이터
  FEVD           — 분산분해 데이터
  VAR계수        — VAR 회귀계수 행렬
"""

import logging

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, size=10)
_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_TITLE_FONT  = Font(bold=True, size=13)
_DESC_FONT   = Font(size=9, italic=True, color="444444")
_SIG_FILL    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_DASH_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, padding: int = 4, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + padding, max_width
        )


def _desc_cell(ws, row: int, col: int, text: str) -> None:
    """차트 설명 텍스트 셀."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _DESC_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# 데이터 시트 빌더 (차트 없음 — 차트는 Dashboard에만)
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, data: dict, results: dict):
    ws = wb.create_sheet("Summary")
    reserves = data["reserves"]
    pearson  = results["pearson"]
    granger  = results["granger"]
    var_res  = results["var"]
    adf_r    = results.get("adf_reserves", {})
    adf_f    = results.get("adf_fx", {})

    def _gstr(direction):
        lag = granger[f"sig_lag_{direction}"]
        p   = granger[f"sig_p_{direction}"]
        return f"lag {lag}개월 유의 (p={p:.4f})" if lag else "유의하지 않음"

    rows = [
        ("항목", "값"),
        ("분석 기간 시작", str(reserves.index[0])),
        ("분석 기간 종료",  str(reserves.index[-1])),
        ("총 관측 월수",   f"{len(reserves)}개월"),
        ("", ""),
        ("외환보유고 ADF p-value", f"{adf_r.get('p_value', 0):.4f}" if adf_r else ""),
        ("환율 ADF p-value",       f"{adf_f.get('p_value', 0):.4f}" if adf_f else ""),
        ("차분 적용", "1차 차분"),
        ("", ""),
        ("피어슨 상관계수 (r)", f"{pearson['r']:.4f}"),
        ("피어슨 p-value",      f"{pearson['p_value']:.4f}"),
        ("피어슨 n",            str(pearson["n"])),
        ("", ""),
        ("Granger 보유고→환율", _gstr("x_to_y")),
        ("Granger 환율→보유고", _gstr("y_to_x")),
        ("", ""),
        ("VAR 최적 lag (AIC)",        f"{var_res['optimal_lag']}개월"),
        ("IRF peak 기간",             f"{var_res['peak_month']}개월 후"),
        ("FEVD 외환보유고 기여(12개월)", f"{var_res['fevd_pct']:.2f}%"),
    ]
    for r_idx, (key, val) in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=key)
        ws.cell(row=r_idx, column=2, value=val)
        if r_idx == 1:
            _style_header(ws, r_idx, 2)
    _auto_width(ws)
    return ws


def _build_timeseries(wb: Workbook, data: dict):
    ws = wb.create_sheet("TimeSeries")
    reserves = data["reserves"]
    usdkrw   = data["usdkrw"]

    headers = ["기간", "외환보유고(억달러)", "원/달러(월말)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for r, period in enumerate(reserves.index, start=2):
        ws.cell(row=r, column=1, value=str(period))
        ws.cell(row=r, column=2, value=round(float(reserves[period]), 2))
        val = usdkrw.get(period, None)
        ws.cell(row=r, column=3, value=round(float(val), 2) if val is not None else None)
    _auto_width(ws)
    return ws


def _build_delta_return(wb: Workbook, data: dict, results: dict):
    ws = wb.create_sheet("DeltaReturn")
    res_delta = data["reserves_delta"]
    fx_ret    = data["fx_return"]
    aligned   = pd.DataFrame({"delta": res_delta, "ret": fx_ret}).dropna()

    x_vals = aligned["delta"].values
    y_vals = aligned["ret"].values
    slope, intercept = np.polyfit(x_vals, y_vals, 1)

    headers = ["기간", "보유고변화(억$)", "환율변동률(%)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for r, (period, row) in enumerate(aligned.iterrows(), start=2):
        ws.cell(row=r, column=1, value=str(period))
        ws.cell(row=r, column=2, value=round(float(row["delta"]), 3))
        ws.cell(row=r, column=3, value=round(float(row["ret"]), 4))

    # 회귀선 2점 (E, F열) — 차트에서 직선 series로 사용
    ws.cell(row=1, column=5, value="회귀_X").font = _HEADER_FONT
    ws.cell(row=1, column=6, value="회귀_Y").font = _HEADER_FONT
    x_min, x_max = float(x_vals.min()), float(x_vals.max())
    ws.cell(row=2, column=5, value=round(x_min, 3))
    ws.cell(row=2, column=6, value=round(slope * x_min + intercept, 4))
    ws.cell(row=3, column=5, value=round(x_max, 3))
    ws.cell(row=3, column=6, value=round(slope * x_max + intercept, 4))

    _auto_width(ws)
    return ws, len(aligned)


def _build_granger(wb: Workbook, results: dict):
    ws  = wb.create_sheet("Granger")
    gr  = results["granger"]
    max_lag = config.GRANGER_MAX_LAG

    # 넓은 포맷 (차트용): Lag | p_xy | p_yx | α=0.05
    wide = ["Lag(개월)", "보유고→환율 p", "환율→보유고 p", "α=0.05"]
    for col, h in enumerate(wide, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(wide))

    for lag in range(1, max_lag + 1):
        r = lag + 1
        p_xy = gr["x_to_y"][lag]
        p_yx = gr["y_to_x"][lag]
        ws.cell(row=r, column=1, value=lag)
        ws.cell(row=r, column=2, value=round(p_xy, 4))
        ws.cell(row=r, column=3, value=round(p_yx, 4))
        ws.cell(row=r, column=4, value=0.05)
        if p_xy < 0.05:
            ws.cell(row=r, column=2).fill = _SIG_FILL
        if p_yx < 0.05:
            ws.cell(row=r, column=3).fill = _SIG_FILL

    # 긴 포맷 (읽기용): F~I열
    long_h = ["Lag", "방향", "p-value", "유의"]
    for col, h in enumerate(long_h, start=6):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, 9)

    r2 = 2
    for lag in range(1, max_lag + 1):
        for direction, p_val in [
            ("보유고 → 환율", gr["x_to_y"][lag]),
            ("환율 → 보유고", gr["y_to_x"][lag]),
        ]:
            ws.cell(row=r2, column=6, value=lag)
            ws.cell(row=r2, column=7, value=direction)
            ws.cell(row=r2, column=8, value=round(p_val, 4))
            ws.cell(row=r2, column=9, value="Y" if p_val < 0.05 else "N")
            if p_val < 0.05:
                ws.cell(row=r2, column=8).fill = _SIG_FILL
            r2 += 1

    _auto_width(ws)
    return ws


def _build_irf(wb: Workbook, results: dict):
    ws      = wb.create_sheet("IRF")
    var_res = results["var"]
    periods = config.IRF_PERIODS

    headers = ["기간(개월)", "IRF", "하한(95%)", "상한(95%)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i in range(periods + 1):
        r = i + 2
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=round(float(var_res["irf_values"][i]), 6))
        ws.cell(row=r, column=3, value=round(float(var_res["irf_lower"][i]),  6))
        ws.cell(row=r, column=4, value=round(float(var_res["irf_upper"][i]),  6))

    _auto_width(ws)
    return ws


def _build_fevd(wb: Workbook, results: dict):
    ws      = wb.create_sheet("FEVD")
    var_res = results["var"]
    fevd_obj = var_res["fevd_obj"]
    from engine.var_model import FX_COL, RESERVES_COL
    cols    = [RESERVES_COL, FX_COL]
    fx_idx  = cols.index(FX_COL)
    res_idx = cols.index(RESERVES_COL)
    periods = config.IRF_PERIODS

    headers = ["기간", "외환보유고 기여(%)", "기타 기여(%)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for t in range(1, periods + 1):
        r       = t + 1
        res_pct = float(fevd_obj.decomp[fx_idx, t - 1, res_idx]) * 100
        # 기간을 문자열로 저장 → Excel이 범주축으로 인식 (숫자 → 빈 라벨 버그 방지)
        ws.cell(row=r, column=1, value=f"{t}개월")
        ws.cell(row=r, column=2, value=round(res_pct, 2))
        ws.cell(row=r, column=3, value=round(100 - res_pct, 2))

    _auto_width(ws)
    return ws


def _build_var_params(wb: Workbook, results: dict):
    ws = wb.create_sheet("VAR계수")
    params_df: pd.DataFrame = results["var"]["var_params"]

    ws.cell(row=1, column=1, value="VAR 모델 계수 행렬").font = _TITLE_FONT
    ws.cell(row=2, column=1, value="변수")
    for col_idx, col_name in enumerate(params_df.columns, start=2):
        ws.cell(row=2, column=col_idx, value=col_name)
    _style_header(ws, 2, len(params_df.columns) + 1)

    for r_idx, (idx_val, row_vals) in enumerate(params_df.iterrows(), start=3):
        ws.cell(row=r_idx, column=1, value=str(idx_val))
        for col_idx, val in enumerate(row_vals, start=2):
            ws.cell(row=r_idx, column=col_idx, value=round(float(val), 6))

    _auto_width(ws)
    return ws


# ---------------------------------------------------------------------------
# Dashboard — 모든 차트 + 설명을 한 시트에
# ---------------------------------------------------------------------------

def _add_label(ws, row: int, col: int, title: str, desc: str) -> None:
    """차트 상단 제목 + 한 줄 설명."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 10)
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = Font(bold=True, size=11)
    cell.fill = _DASH_FILL

    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 10)
    desc_cell = ws.cell(row=row + 1, column=col, value=desc)
    desc_cell.font = _DESC_FONT
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row + 1].height = 28


def _sheet_dashboard(
    wb, data, results,
    ws_ts, ws_dr, n_dr, ws_gr, ws_irf, ws_fevd,
) -> None:
    ws = wb.create_sheet("Dashboard")
    pearson  = results["pearson"]
    var_res  = results["var"]
    granger  = results["granger"]
    max_lag  = config.GRANGER_MAX_LAG
    periods  = config.IRF_PERIODS
    n_ts     = len(data["reserves"])

    # ── 페이지 제목 ──────────────────────────────────────────────
    ws.merge_cells("A1:Z1")
    reserves = data["reserves"]
    start_label = str(reserves.index[0])
    end_label   = str(reserves.index[-1])
    n_months    = len(reserves)
    title_cell = ws.cell(row=1, column=1,
        value=f"외환보유고 ↔ 원/달러 환율 관계 분석  |  {start_label} ~ {end_label}  ({n_months}개월)")
    title_cell.font  = Font(bold=True, size=14)
    title_cell.fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── 차트 1: 시계열 (전폭) ────────────────────────────────────
    CHART_ROW_1 = 3
    _add_label(ws, CHART_ROW_1, 1,
        "① 외환보유고 & 원/달러 환율 시계열",
        f"외환보유고(좌축·파랑)와 원/달러 월말 환율(우축·빨강). "
        f"IMF 위기(1997), GFC(2008), 코로나(2020), 킹달러(2022) 구간에서 보유고 감소 + 환율 급등 패턴이 반복됨.")

    c1 = LineChart()
    c1.title  = None
    c1.style  = 10
    c1.height = 14
    c1.width  = 42
    c1.y_axis.title = "외환보유고 (억달러)"
    c1.y_axis.axId  = 200

    data_res = Reference(ws_ts, min_col=2, min_row=1, max_row=n_ts + 1)
    c1.add_data(data_res, titles_from_data=True)
    cats_ts = Reference(ws_ts, min_col=1, min_row=2, max_row=n_ts + 1)
    c1.set_categories(cats_ts)

    c1_fx = LineChart()
    c1_fx.y_axis.axId   = 100
    c1_fx.y_axis.crosses = "max"
    c1_fx.y_axis.title  = "원/달러 환율"
    data_fx = Reference(ws_ts, min_col=3, min_row=1, max_row=n_ts + 1)
    c1_fx.add_data(data_fx, titles_from_data=True)
    c1_fx.set_categories(cats_ts)
    c1 += c1_fx

    c1.series[0].graphicalProperties.line.solidFill = "4472C4"
    c1.series[0].graphicalProperties.line.width     = 12000
    if len(c1.series) > 1:
        c1.series[1].graphicalProperties.line.solidFill = "C00000"
        c1.series[1].graphicalProperties.line.width     = 10000

    ws.add_chart(c1, f"A{CHART_ROW_1 + 2}")

    # ── 차트 2 & 3 나란히 ────────────────────────────────────────
    CHART_ROW_2 = 35
    _add_label(ws, CHART_ROW_2, 1,
        "② 산점도: 보유고 변화량 vs 환율 변동률",
        f"피어슨 r = {pearson['r']:.3f} (p<0.001, n={pearson['n']}). "
        f"보유고 증가 시 환율 하락 경향. 회귀선(빨강).")
    _add_label(ws, CHART_ROW_2, 14,
        "③ Granger 인과성 p-value by Lag",
        f"α=0.05(점선) 하방이 유의. "
        f"보유고→환율: lag {granger['sig_lag_x_to_y']}개월(p={granger['sig_p_x_to_y']:.3f}), "
        f"환율→보유고: lag {granger['sig_lag_y_to_x']}개월(p={granger['sig_p_y_to_x']:.3f}).")

    # Chart 2: 산점도
    sc = ScatterChart()
    sc.title  = None
    sc.style  = 13
    sc.height = 14
    sc.width  = 20
    sc.x_axis.title = "보유고 변화량 (억달러)"
    sc.y_axis.title = "환율 변동률 (%)"

    xvals = Reference(ws_dr, min_col=2, min_row=2, max_row=n_dr + 1)
    yvals = Reference(ws_dr, min_col=3, min_row=2, max_row=n_dr + 1)
    s_data = Series(yvals, xvals, title="월별 관측")
    s_data.marker.symbol = "circle"
    s_data.marker.size   = 3
    s_data.graphicalProperties.line.noFill = True
    s_data.marker.graphicalProperties.solidFill = "4472C4"
    s_data.marker.graphicalProperties.line.solidFill = "4472C4"
    sc.series.append(s_data)

    rx = Reference(ws_dr, min_col=5, min_row=2, max_row=3)
    ry = Reference(ws_dr, min_col=6, min_row=2, max_row=3)
    s_reg = Series(ry, rx, title="회귀선")
    s_reg.graphicalProperties.line.solidFill = "C00000"
    s_reg.graphicalProperties.line.width     = 20000
    sc.series.append(s_reg)

    ws.add_chart(sc, f"A{CHART_ROW_2 + 2}")

    # Chart 3: Granger
    lc3 = LineChart()
    lc3.title  = None
    lc3.style  = 10
    lc3.height = 14
    lc3.width  = 20
    lc3.y_axis.title = "p-value"
    lc3.x_axis.title = "Lag (개월)"
    lc3.y_axis.scaling.min = 0.0
    lc3.y_axis.scaling.max = 1.0

    # 3개 시리즈: p_xy, p_yx, α=0.05
    g_data = Reference(ws_gr, min_col=2, max_col=4, min_row=1, max_row=max_lag + 1)
    lc3.add_data(g_data, titles_from_data=True)
    cats_gr = Reference(ws_gr, min_col=1, min_row=2, max_row=max_lag + 1)
    lc3.set_categories(cats_gr)

    lc3.series[0].graphicalProperties.line.solidFill = "4472C4"
    lc3.series[1].graphicalProperties.line.solidFill = "C00000"
    lc3.series[1].graphicalProperties.line.dashDot   = "dash"
    lc3.series[2].graphicalProperties.line.solidFill = "000000"
    lc3.series[2].graphicalProperties.line.dashDot   = "dot"
    lc3.series[2].graphicalProperties.line.width     = 15000

    ws.add_chart(lc3, f"N{CHART_ROW_2 + 2}")

    # ── 차트 4 & 5 나란히 ────────────────────────────────────────
    CHART_ROW_3 = 65
    _add_label(ws, CHART_ROW_3, 1,
        f"④ VAR IRF: 외환보유고 충격 → 환율 반응  (VAR lag={var_res['optimal_lag']})",
        f"보유고 1단위 충격 후 환율 반응 경로. "
        f"파선=95% 신뢰구간. Peak {var_res['peak_month']}개월 후 최대 반응 후 수렴.")
    _add_label(ws, CHART_ROW_3, 14,
        f"⑤ FEVD: 환율 분산 분해  (12개월 기여 {var_res['fevd_pct']:.1f}%)",
        f"환율 변동성 중 외환보유고(파랑)가 설명하는 비율. "
        f"12개월 기준 {var_res['fevd_pct']:.1f}%, 나머지 {100-var_res['fevd_pct']:.1f}%는 외부 요인.")

    # Chart 4: IRF
    n_irf = periods + 1
    lc4 = LineChart()
    lc4.title  = None
    lc4.style  = 10
    lc4.height = 14
    lc4.width  = 20
    lc4.y_axis.title = "환율 반응"
    lc4.x_axis.title = "기간 (개월)"

    irf_data = Reference(ws_irf, min_col=2, max_col=4, min_row=1, max_row=n_irf + 1)
    lc4.add_data(irf_data, titles_from_data=True)
    cats_irf = Reference(ws_irf, min_col=1, min_row=2, max_row=n_irf + 1)
    lc4.set_categories(cats_irf)

    lc4.series[0].graphicalProperties.line.solidFill = "4472C4"
    lc4.series[0].graphicalProperties.line.width     = 25000
    lc4.series[1].graphicalProperties.line.solidFill = "A0A0A0"
    lc4.series[1].graphicalProperties.line.dashDot   = "dash"
    lc4.series[2].graphicalProperties.line.solidFill = "A0A0A0"
    lc4.series[2].graphicalProperties.line.dashDot   = "dash"

    ws.add_chart(lc4, f"A{CHART_ROW_3 + 2}")

    # Chart 5: FEVD
    bc5 = BarChart()
    bc5.type     = "col"
    bc5.grouping = "stacked"
    bc5.overlap  = 100
    bc5.title    = None
    bc5.style    = 10
    bc5.height   = 14
    bc5.width    = 20
    bc5.y_axis.title = "기여 비율 (%)"
    bc5.x_axis.title = "기간 (개월)"
    bc5.y_axis.scaling.min = 0
    bc5.y_axis.scaling.max = 100

    fevd_data = Reference(ws_fevd, min_col=2, max_col=3, min_row=1, max_row=periods + 1)
    bc5.add_data(fevd_data, titles_from_data=True)
    cats_fevd = Reference(ws_fevd, min_col=1, min_row=2, max_row=periods + 1)
    bc5.set_categories(cats_fevd)

    bc5.series[0].graphicalProperties.solidFill = "4472C4"
    bc5.series[1].graphicalProperties.solidFill = "D3D3D3"

    ws.add_chart(bc5, f"N{CHART_ROW_3 + 2}")


# ---------------------------------------------------------------------------
# 메인 진입점
# ---------------------------------------------------------------------------

def generate_excel(
    data: dict,
    results: dict,
    chart_path: str | None = None,
    output_path: str | None = None,
) -> str:
    if output_path is None:
        output_path = str(config.EXCEL_PATH)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 데이터 시트 (Dashboard에서 참조하므로 먼저 생성)
    ws_summary = _build_summary(wb, data, results)
    ws_ts      = _build_timeseries(wb, data)
    ws_dr, n_dr = _build_delta_return(wb, data, results)
    ws_gr      = _build_granger(wb, results)
    ws_irf     = _build_irf(wb, results)
    ws_fevd    = _build_fevd(wb, results)
    ws_var     = _build_var_params(wb, results)

    # Dashboard (맨 앞으로 이동)
    _sheet_dashboard(wb, data, results, ws_ts, ws_dr, n_dr, ws_gr, ws_irf, ws_fevd)
    wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames) - 1))

    wb.save(output_path)
    logger.info("Excel 저장 완료 (Dashboard 통합): %s", output_path)
    return output_path
